from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory
from database import MySQLDB, RedshiftDB
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv, find_dotenv
import click
import os
import csv
import openai
import traceback
from openai import OpenAI


#################################################


# Set OpenAI API Key
api_key = os.getenv("OPENAI_API_KEY")

# Set the API key
client = OpenAI(api_key=api_key)

if not api_key:
    print("WARNING: OPENAI_API_KEY not found in environment variables!")
else:
    print("OpenAI API key loaded successfully")


#################################################


app = Flask(__name__)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'
app.secret_key = 'LeatherShoes'


class User(UserMixin):
    def __init__(self, user_id, username, password_hash):
        self.id = user_id
        self.username = username
        self.password_hash = password_hash

    @staticmethod
    def create_user(username, password):
        """Create a new user, and store the encrypted password in the database."""
        password_hash = generate_password_hash(password)
        with MySQLDB("leathershoesuser") as db:
            db.execute_query("INSERT INTO users (username, password_hash) VALUES (%s, %s)", (username, password_hash))
            db.commit()

    def validate_password(self, password):
        """Verify the user's entered password."""
        return check_password_hash(self.password_hash, password)
    
    @staticmethod
    def get_user(username):
        with MySQLDB("leathershoesuser") as db:
            result = db.execute_query("SELECT id, username, password_hash FROM users WHERE username = %s", (username,))
            if result:
                return User(*result[0])
        return None


#################################################


app.config['UPLOAD_DIRECTORY'] ="uploads"
app.config['REPORT_FOLDER'] ="reports"


@login_manager.user_loader
def load_user(user_id):
    with MySQLDB("leathershoesuser") as db:
        result = db.execute_query("SELECT id, username, password_hash FROM users WHERE id = %s", (int(user_id),))
        if result:
            return User(*result[0])
    return None


@app.route('/')
@app.route('/home')
def home():
    return render_template('home.html', page_title="首頁")

@app.route('/dashboard')
def dashboard():
    with RedshiftDB() as redshift_db:
        cursor = redshift_db.cursor

        try:
               
            # 1. Total Profit from Sales Last Month (POS System)
            cursor.execute("""
                SELECT D.yearnumber, D.monthnumber, SUM(S.Profit) AS TotalProfit
                FROM Fact_Sales S
                JOIN Dim_Date D ON S.OrderDateID = D.DateID
                WHERE ( D.yearnumber = DATE_PART(YEAR, CURRENT_DATE)
                        AND D.monthnumber = DATE_PART(MONTH, CURRENT_DATE) - 1 AND S.SalesCategoryID IN (2,3) )
                OR
                    ( DATE_PART(MONTH, CURRENT_DATE) = 1
                        AND D.yearnumber = DATE_PART(YEAR, CURRENT_DATE) - 1
                        AND D.monthnumber = 12 AND S.SalesCategoryID IN (2,3) )
                GROUP BY D.yearnumber, D.monthnumber;
            """)
            TotalProfitLastMonth = cursor.fetchone()[2]


            # 2. Branch-wise Profit Last Month vs Previous Month
            cursor.execute("""
                SELECT 
                    B.BranchName, 
                    COALESCE(SUM(CASE WHEN DATE_TRUNC('month', D.FullDate) = DATE_TRUNC('month', CURRENT_DATE) - INTERVAL '1 month' THEN S.Profit END), 0) AS LastMonthProfit,
                    COALESCE(SUM(CASE WHEN DATE_TRUNC('month', D.FullDate) = DATE_TRUNC('month', CURRENT_DATE) - INTERVAL '2 months' THEN S.Profit END), 0) AS PrevMonthProfit
                FROM Fact_Sales S
                JOIN Dim_Branch B ON S.SoldBranchID = B.BranchID
                JOIN Dim_Date D ON S.OrderDateID = D.DateID
                WHERE D.FullDate >= CURRENT_DATE - INTERVAL '4 months' AND S.SalesCategoryID IN (2,3)
                GROUP BY B.BranchName
                ORDER BY LastMonthProfit DESC;
            """)        
            BranchProfitLastMonth = cursor.fetchall()


            # 3. Branch-wise Profit This Year vs Last Year
            cursor.execute("""
                SELECT
                    B.BranchName,
                    COALESCE(SUM(CASE WHEN D.YearNumber = EXTRACT(YEAR FROM CURRENT_DATE) 
                           AND D.MonthNumber <= EXTRACT(MONTH FROM CURRENT_DATE) THEN S.Profit END), 0) AS ThisYearProfit,
                    COALESCE(SUM(CASE WHEN D.YearNumber = EXTRACT(YEAR FROM CURRENT_DATE) - 1
                            AND D.MonthNumber <= EXTRACT(MONTH FROM CURRENT_DATE) THEN S.Profit END), 0) AS LastYearProfit
                FROM Fact_Sales S
                JOIN Dim_Branch B ON S.SoldBranchID = B.BranchID
                JOIN Dim_Date D ON S.OrderDateID = D.DateID
                WHERE S.SalesCategoryID IN (2,3)
                GROUP BY B.BranchName
                HAVING
                    COALESCE(SUM(CASE WHEN D.YearNumber = EXTRACT(YEAR FROM CURRENT_DATE) 
                           AND D.MonthNumber <= EXTRACT(MONTH FROM CURRENT_DATE) THEN S.Profit END), 0) > 0
                    AND COALESCE(SUM(CASE WHEN d.YearNumber = EXTRACT(YEAR FROM CURRENT_DATE) - 1 
                           AND D.MonthNumber <= EXTRACT(MONTH FROM CURRENT_DATE) THEN S.Profit END), 0) > 0
                ORDER BY
                    ThisYearProfit DESC;
            """)
            BranchProfitThisYear = cursor.fetchall()


            # 4️. Supplier-wise Profit Last Month
            cursor.execute("""
                SELECT 
                    SP.SupplierName, 
                    COALESCE(SUM(S.Profit), 0) AS LastMonthProfit
                FROM Fact_Sales S
                JOIN Dim_Supplier SP ON S.SupplierID = SP.SupplierID
                JOIN Dim_Date D ON S.OrderDateID = D.DateID
                WHERE DATE_TRUNC('month', D.FullDate) = DATE_TRUNC('month', CURRENT_DATE) - INTERVAL '1 month' 
                           AND S.SalesCategoryID IN (2,3)
                GROUP BY SP.SupplierName
                ORDER BY LastMonthProfit DESC
                LIMIT 15;
            """)
            SupplierProfitLastMonth = cursor.fetchall()


            # 5️. Supplier-wise Profit This Year
            cursor.execute("""
                SELECT 
                    SP.SupplierName, 
                    COALESCE(SUM(S.Profit), 0) AS ThisYearProfit
                FROM Fact_Sales S
                JOIN Dim_Supplier SP ON S.SupplierID = SP.SupplierID
                JOIN Dim_Date D ON S.OrderDateID = D.DateID
                WHERE D.YearNumber = EXTRACT(YEAR FROM CURRENT_DATE) AND S.SalesCategoryID IN (2,3)
                GROUP BY SP.SupplierName
                ORDER BY ThisYearProfit DESC
                LIMIT 15;
            """)
            SupplierProfitThisYear = cursor.fetchall()


            # 6️. Top Suppliers by Total Profit
            cursor.execute("""
                SELECT 
                    SP.SupplierName, 
                    COALESCE(SUM(S.Profit), 0) AS TotalProfit
                FROM Fact_Sales S
                JOIN Dim_Supplier SP ON S.SupplierID = SP.SupplierID
                WHERE S.SalesCategoryID IN (2,3)
                GROUP BY SP.SupplierName
                ORDER BY TotalProfit DESC
                LIMIT 15;
            """)
            SupplierTotalProfit = cursor.fetchall()


            # Prepare chart data for visualization
            chart_data = {
                'TotalProfitLastMonth': TotalProfitLastMonth,
                'BranchProfitLastMonth': {
                    'labels': [row[0] for row in BranchProfitLastMonth],
                    'LastMonthProfit': [row[1] for row in BranchProfitLastMonth],
                    'PrevMonthProfit': [row[2] for row in BranchProfitLastMonth]
                },
                'BranchProfitThisYear': {
                    'labels': [row[0] for row in BranchProfitThisYear],
                    'ThisYearProfit': [row[1] for row in BranchProfitThisYear],
                    'LastYearProfit': [row[2] for row in BranchProfitThisYear]
                },
                'SupplierProfitLastMonth': {
                    'labels': [row[0] for row in SupplierProfitLastMonth],
                    'profit': [row[1] for row in SupplierProfitLastMonth]
                },
                'SupplierProfitThisYear': {
                    'labels': [row[0] for row in SupplierProfitThisYear],
                    'profit': [row[1] for row in SupplierProfitThisYear]
                },
                'SupplierTotalProfit': {
                    'labels': [row[0] for row in SupplierTotalProfit],
                    'profit': [row[1] for row in SupplierTotalProfit]
                }
            }

        except Exception as e:
            print(f"Error executing queries: {e}")
            return render_template('error.html', error_message=str(e))

    return render_template('dashboard.html', chart_data=chart_data, page_title="Dashboard")


@app.route('/accounting', methods=["GET", "POST"])
@login_required
def accounting():
    if request.method == "POST":
        file = request.files["datafile"]

        if file:
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_DIRECTORY'], filename)
            file.save(file_path)
            flash("File Uploaded Successfully")

            data = []
            with open(file_path) as f:
                reader_obj = csv.reader(f)

                next(reader_obj)  # Skip the header row
                for row in reader_obj:
                    data.append(tuple(row))
        
        if data:
            with MySQLDB() as db:
                try:
                    # Insert data into the database
                    cursor = db.cursor
                    cursor.executemany("INSERT INTO accounting (BranchID, EmployeeID, Date, AccountingTypeID, Description, Amount) VALUES (%s, %s, %s, %s, %s, %s)", data)
                    db.commit()
                    flash("Data Imported Successfully")
                except Exception as e:
                    print(f"Error inserting data: {e}")
                    flash("Data Import Failed")
        
        return redirect(url_for("accounting"))

    # Fetch Accounting Data from MySQL RDS
    with MySQLDB() as mysql_db:
        cursor = mysql_db.cursor
        try:
            # 1. Branch-wise Profit Last Month vs Previous Month
            cursor.execute("""
                SELECT 
                    BranchID, 
                    COALESCE(SUM(CASE WHEN DATE_FORMAT(Date, '%Y-%m') = DATE_FORMAT(DATE_SUB(CURDATE(), INTERVAL 1 MONTH), '%Y-%m') THEN Amount END), 0) AS LastMonthProfit,
                    COALESCE(SUM(CASE WHEN DATE_FORMAT(Date, '%Y-%m') = DATE_FORMAT(DATE_SUB(CURDATE(), INTERVAL 2 MONTH), '%Y-%m') THEN Amount END), 0) AS PrevMonthProfit
                FROM accounting
                GROUP BY BranchID
                ORDER BY LastMonthProfit DESC;
            """)
            BranchProfitLastMonth = cursor.fetchall()

            # 2. Branch-wise Profit This Year vs Last Year
            cursor.execute("""
                SELECT 
                    BranchID, 
                    COALESCE(SUM(CASE WHEN YEAR(Date) = YEAR(CURDATE()) THEN Amount END), 0) AS ThisYearProfit,
                    COALESCE(SUM(CASE WHEN YEAR(Date) = YEAR(CURDATE()) - 1 THEN Amount END), 0) AS LastYearProfit
                FROM accounting
                GROUP BY BranchID
                ORDER BY ThisYearProfit DESC;
            """)
            BranchProfitThisYear = cursor.fetchall()

            # 3. Top 10 Expenses Last Month vs Previous Month
            cursor.execute("""
                SELECT 
                    BranchID, 
                    AccountingTypeID, 
                    COALESCE(SUM(CASE WHEN DATE_FORMAT(Date, '%Y-%m') = DATE_FORMAT(DATE_SUB(CURDATE(), INTERVAL 1 MONTH), '%Y-%m') THEN Amount END), 0) AS LastMonthExpense,
                    COALESCE(SUM(CASE WHEN DATE_FORMAT(Date, '%Y-%m') = DATE_FORMAT(DATE_SUB(CURDATE(), INTERVAL 2 MONTH), '%Y-%m') THEN Amount END), 0) AS PrevMonthExpense
                FROM accounting
                WHERE Amount < 0  -- Only Expenses
                GROUP BY BranchID, AccountingTypeID
                ORDER BY LastMonthExpense ASC
                LIMIT 10;
            """)
            TopExpensesLastMonth = cursor.fetchall()

            # 4. Top 10 Expenses This Year vs Last Year
            cursor.execute("""
                SELECT 
                    BranchID, 
                    AccountingTypeID, 
                    COALESCE(SUM(CASE WHEN YEAR(Date) = YEAR(CURDATE()) THEN Amount END), 0) AS ThisYearExpense,
                    COALESCE(SUM(CASE WHEN YEAR(Date) = YEAR(CURDATE()) - 1 THEN Amount END), 0) AS LastYearExpense
                FROM accounting
                WHERE Amount < 0  -- Only Expenses
                GROUP BY BranchID, AccountingTypeID
                ORDER BY ThisYearExpense ASC
                LIMIT 10;
            """)
            TopExpensesThisYear = cursor.fetchall()

        except Exception as e:
            print(f"Error executing accounting queries: {e}")

    # Prepare chart data for visualization
    chart_data = {
        'BranchProfitLastMonth': {
            'labels': [row[0] for row in BranchProfitLastMonth],  # BranchID
            'LastMonthProfit': [row[1] for row in BranchProfitLastMonth],
            'PrevMonthProfit': [row[2] for row in BranchProfitLastMonth]
        },
        'BranchProfitThisYear': {
            'labels': [row[0] for row in BranchProfitThisYear],  # BranchID
            'ThisYearProfit': [row[1] for row in BranchProfitThisYear],
            'LastYearProfit': [row[2] for row in BranchProfitThisYear]
        },
        'TopExpensesLastMonth': {
            'labels': [f"Branch {row[0]} - Type {row[1]}" for row in TopExpensesLastMonth],
            'LastMonthExpense': [row[2] for row in TopExpensesLastMonth],
            'PrevMonthExpense': [row[3] for row in TopExpensesLastMonth]
        },
        'TopExpensesThisYear': {
            'labels': [f"Branch {row[0]} - Type {row[1]}" for row in TopExpensesThisYear],
            'ThisYearExpense': [row[2] for row in TopExpensesThisYear],
            'LastYearExpense': [row[3] for row in TopExpensesThisYear]
        }
    }

    with MySQLDB() as db:
        try:
            # Get accounting data from the database
            cursor = db.cursor
            cursor.execute("SELECT A.AccountingID, B.Name, E.Name, A.Date, AC.AccountingType, A.Description, A.Amount FROM accounting A JOIN branch B ON A.BranchID = B.BranchID JOIN employee E ON A.EmployeeID = E.EmployeeID JOIN accountingtype AC ON A.AccountingTypeID = AC.AccountingTypeID ORDER BY A.Date DESC")
            accounting_data = cursor.fetchall()
        except Exception as e:
            print(f"Error fetching data: {e}")
            return render_template('error.html', error_message=str(e))

    return render_template("accounting.html", chart_data=chart_data , accounting_data=accounting_data, page_title="Accounting")


@app.route('/accounting_add', methods=["GET", "POST"])
@login_required
def accounting_add():
    if request.method == "POST":
        data = [
            request.form["BranchID"],
            request.form["EmployeeID"],
            request.form["Date"],
            request.form["AccountingTypeID"],
            request.form["Description"],
            request.form["Amount"]
        ]

        with MySQLDB() as db:
            try:
                # Insert data into the database
                db.execute_query("INSERT INTO accounting (BranchID, EmployeeID, Date, AccountingTypeID, Description, Amount) VALUES (%s, %s, %s, %s, %s, %s)", data)
                db.commit()
                flash("Data Inserted Successfully")
            except Exception as e:
                print(f"Error inserting data: {e}")
                flash("Data Insert Failed")
        
        return redirect(url_for("accounting"))

    return render_template("accounting_add.html", page_title="Create Accounting Data")


@app.route('/accounting_edit/<id>', methods=["GET", "POST"])
@login_required
def accounting_edit(id):
    print(f"In accounting_edit: {id}")

    if request.method == "POST":
        data = [
            request.form["BranchID"],
            request.form["EmployeeID"],
            request.form["Date"],
            request.form["AccountingTypeID"],
            request.form["Description"],
            request.form["Amount"],
            id
        ]

        with MySQLDB() as db:
            try:
                # Update data in the database
                db.execute_query("UPDATE accounting SET BranchID = %s, EmployeeID = %s, Date = %s, AccountingTypeID = %s, Description = %s, Amount = %s WHERE AccountingID = %s", data)
                db.commit()
                flash("Data Updated Successfully")
            except Exception as e:
                print(f"Error updating data: {e}")
                flash("Data Update Failed")
        
        return redirect(url_for("accounting"))

    with MySQLDB() as db:
        try:
            # Get accounting data from the database
            accounting_id = id
            cursor = db.cursor
            cursor.execute("SELECT * FROM accounting WHERE AccountingID = %s", (accounting_id,))
            accounting_data = cursor.fetchone()
        except Exception as e:
            print(f"Error fetching data: {e}")
            return render_template('error.html', error_message=str(e))

    return render_template("accounting_edit.html", accounting_data=accounting_data, page_title="Edit Accounting Data")


@app.route('/accounting_delete/<id>',methods=["GET","POST"])
@login_required
def accounting_delete(id):
    print(f"In accounting_delete: {id}")

    with MySQLDB() as db:
        try:
            # Delete data from the database
            accounting_id = id
            db.execute_query("DELETE FROM accounting WHERE AccountingID = %s", (accounting_id,))
            db.commit()
            flash("Data Deleted Successfully")
        except Exception as e:
            print(f"Error deleting data: {e}")
            flash("Data Delete Failed")

    return redirect(url_for("accounting"))


@app.route('/accounting_download')
def accounting_download():
    wb = Workbook()
    ws = wb.active
    with MySQLDB() as db:
        try:
            # Get accounting data from the database
            cursor = db.cursor
            cursor.execute("SELECT A.AccountingID, B.Name, E.Name, A.Date, AC.AccountingType, A.Description, A.Amount FROM accounting A JOIN branch B ON A.BranchID = B.BranchID JOIN employee E ON A.EmployeeID = E.EmployeeID JOIN accountingtype AC ON A.AccountingTypeID = AC.AccountingTypeID ORDER BY A.Date DESC")
            accounting_data = cursor.fetchall()
        except Exception as e:
            print(f"Error fetching data: {e}")
            return render_template('error.html', error_message=str(e))

    # Write data to the worksheet
    #Headers
    row = 1
    ws.cell(row=row, column=1).value = "Branch"
    ws.cell(row=row, column=2).value = "Employee"
    ws.cell(row=row, column=3).value = "Date"
    ws.cell(row=row, column=4).value = "AccountingType"
    ws.cell(row=row, column=5).value = "Description"
    ws.cell(row=row, column=6).value = "Amount"

    for col in range(1, 7):
        ws.cell(row=row, column=col).style = "Accent1"

    #Data
    row = 2
    for data_row in accounting_data:
        ws.cell(row=row, column=1).value = data_row[1]
        ws.cell(row=row, column=2).value = data_row[2]
        ws.cell(row=row, column=3).value = data_row[3]
        ws.cell(row=row, column=4).value = data_row[4]
        ws.cell(row=row, column=5).value = data_row[5]
        ws.cell(row=row, column=6).value = data_row[6]
        row += 1

    for i in range(1,ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True
        ws.column_dimensions[get_column_letter(i)].auto_size = True

    file_name="accounting_data.xlsx"
    wb.save(app.config['REPORT_FOLDER']+"/"+file_name)


    return send_from_directory(app.config['REPORT_FOLDER'],file_name,as_attachment=True)


@app.route('/explain_chart', methods=["POST"])
def explain_chart():
    try:
        data = request.get_json()

        chart_type = data.get("chart_type", "bar")
        chart_data = data.get("chart_data", {})

        if not chart_data:
            return {"error": "No chart data provided"}, 400

        prompt = f"""
        This is a {chart_type} chart generated from the data of a famous leather shoe company, and I am the general manager and owner of this company. 
        Please explain in detail the meaning of the data on the chart, what the future impact might be, and what the boss needs to pay attention to in the future.
        """

        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", 
                 "content": """You are a data analyst specializing in the leather shoe manufacturing and retail distribution industry in Taiwan. 
                 Provide clear and concise insights, highlighting patterns while avoiding explicit causal statements."""},
                {"role": "user", 
                 "content": prompt}
            ],
            max_tokens=300
        )
        print(response)

        explanation = response.choices[0].message.content
        return {"explanation": explanation}, 200

    except Exception as e:
        error_message = f"Error in /explain_chart: {str(e)}\n{traceback.format_exc()}"
        print(error_message)  # Print error in terminal for debugging
        return {"error": "Internal Server Error. Check logs for details."}, 500



@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        user = User.get_user(username)
        if user and user.validate_password(password):
            login_user(user)
            flash("Login successful.")
            return redirect(url_for("dashboard"))
        flash("Invalid username or password")
    return render_template("login.html")


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash("Logged out successfully.")
    return redirect(url_for('login'))


@app.route('/admin')
@login_required
def admin():
    return "This is the admin page"


if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=True)